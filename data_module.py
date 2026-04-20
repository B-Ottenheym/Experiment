import streamlit as st
import pandas as pd
import numpy as np
from scipy.stats import truncnorm
from scipy.special import expit
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from dictionaries import binary_vars, binary_directionality, categorical_vars, categorical_mappings, continuous_vars, continuous_directionality, integer_continuous_vars, var_groups, group_weights, inter_group_effects, within_group_loadings, var_labels_dutch, var_units_dutch

def generate_synthetic_data(n_projects: int, random_seed: int = 31):

    np.random.seed(random_seed)

    def sample_truncnorm(mu, sigma, a, b, size):
        a_, b_ = (a - mu) / sigma, (b - mu) / sigma
        return truncnorm.rvs(a_, b_, loc=mu, scale=sigma, size=size)

    X = {}

    for v, p in binary_vars.items():
        X[v] = np.random.binomial(1, p, n_projects)

    for v, (cats, probs) in categorical_vars.items():
        X[v] = np.random.choice(cats, size=n_projects, p=probs)

    for v, (mu, sigma, a, b) in continuous_vars.items():
        values = sample_truncnorm(mu, sigma, a, b, n_projects)

        if v in integer_continuous_vars:
            values = np.round(values).astype(int)

        X[v] = values

    df = pd.DataFrame(X)

    df_raw = df.copy()

    df_numerical = df_raw.select_dtypes(include=[np.number])
    df_categorical = df_raw.select_dtypes(exclude=[np.number])

    df_norm = df_raw.copy()

    categorical_columns = df_categorical.columns.tolist()

    for col, mapping in categorical_mappings.items():
        df_norm[col] = df_norm[col].map(mapping)

    for col in df_norm.columns:
        if (
            pd.api.types.is_numeric_dtype(df_norm[col])
            and col not in categorical_columns
        ):
            mn, mx = df_norm[col].min(), df_norm[col].max()

            if mx > mn:
                norm = (df_norm[col] - mn) / (mx - mn)
            else:
                norm = 0.5

            if col in continuous_directionality:
                if continuous_directionality[col] == -1:
                    norm = 1 - norm
            elif col in binary_directionality:
                if binary_directionality[col] == -1:
                    norm = 1 - norm

            df_norm[col] = norm

    G = {}
    for g, vars_ in var_groups.items():
        if g == "context":
            continue
        weights = np.array([within_group_loadings[v] for v in vars_])
        values = df_norm[vars_].values
        G[g] = (values @ weights) / weights.sum()

    G = pd.DataFrame(G)

    G_star = G.copy()
    for g, parents in inter_group_effects.items():
        adjustment = sum(coeff * (G[p] - 0.5) for p, coeff in parents.items())
        G_star[g] = expit(G[g] + adjustment)

    S = sum(group_weights[g] * G_star[g] for g in group_weights)
    R = S.rank(pct=True).values

    tau = 0.47
    alpha = 1.10 / (1 - tau)

    df["is_delayed"] = (R >= tau).astype(int)
    df["delay_pct"] = np.maximum(0, alpha * (R - tau))
    df["delay_days"] = np.ceil(df["delay_pct"] * df["planned_project_duration_days"]).astype(int)

    return df, df_numerical, df_categorical, categorical_vars

def generate_template_excel(
    file_name="project_data_training_template.xlsx",
    n_example_rows=1
):

    numerical_cols = list(continuous_vars.keys())
    categorical_cols = (list(categorical_vars.keys())+ list(binary_vars.keys()))
    target_cols = ["is_delayed", "delay_pct", "delay_days"]

    categorical_options = {}
    for var in binary_vars.keys():
        categorical_options[var] = [0, 1]
    for var, (cats, _probs) in categorical_vars.items():
        categorical_options[var] = list(cats)

    all_cols = numerical_cols + categorical_cols + target_cols
    df_template = pd.DataFrame({col: [] for col in all_cols})

    if n_example_rows > 0:
        example = {col: "" for col in all_cols}
        example["is_delayed"] = 0
        example["delay_pct"] = 0.0
        df_template = pd.concat([df_template, pd.DataFrame([example])], ignore_index=True)

    df_template.to_excel(file_name, index=False)

    wb = load_workbook(file_name)
    ws = wb.active

    for col_idx, col_name in enumerate(df_template.columns, start=1):
        if col_name in categorical_options:
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(map(str, categorical_options[col_name]))}"',
                allow_blank=True
            )
            ws.add_data_validation(dv)
            dv.add(f"{ws.cell(row=2, column=col_idx).coordinate}:{ws.cell(row=500, column=col_idx).coordinate}")

        max_length = max(len(col_name), 12)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 2

    wb.save(file_name)
    return file_name, categorical_options
